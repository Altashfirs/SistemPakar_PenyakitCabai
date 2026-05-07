"""Generate Excel report for diagnosis results."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from data.knowledge_base import SYMPTOMS
from data.recommendations import RECOMMENDATIONS
from engine.cf_engine import build_detail_rows, build_ranking_rows, build_rule_rows


def _write_sheet(ws, title: str, headers: list[str], rows: list[list]):
    ws.title = title
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 40)


def generate_excel_report(results: dict[str, dict], user_inputs: dict[str, float]) -> bytes:
    """Create diagnosis report workbook as bytes."""
    workbook = Workbook()
    summary_sheet = workbook.active

    top_result = next(iter(results.values()))
    _write_sheet(
        summary_sheet,
        "Ringkasan",
        ["Field", "Nilai"],
        [
            ["Penyakit Utama", top_result["name"]],
            ["Kode Penyakit", top_result["code"]],
            ["Nilai CF", top_result["cf_combined"]],
            ["Tingkat Kepercayaan (%)", top_result["percentage"]],
            ["Interpretasi", top_result["label"]],
            ["Rekomendasi Utama", " | ".join(RECOMMENDATIONS.get(top_result["code"], []))],
        ],
    )

    ranking_rows = build_ranking_rows(results)
    ranking_sheet = workbook.create_sheet()
    _write_sheet(
        ranking_sheet,
        "Ranking",
        list(ranking_rows[0].keys()),
        [list(row.values()) for row in ranking_rows],
    )

    input_rows = [
        [code, SYMPTOMS[code], value]
        for code, value in user_inputs.items()
        if value > 0
    ] or [["-", "Tidak ada gejala dipilih", 0.0]]
    input_sheet = workbook.create_sheet()
    _write_sheet(input_sheet, "Input Gejala", ["Kode", "Gejala", "CF User"], input_rows)

    detail_rows = build_detail_rows(results)
    detail_sheet = workbook.create_sheet()
    if detail_rows:
        _write_sheet(
            detail_sheet,
            "Detail CF",
            list(detail_rows[0].keys()),
            [list(row.values()) for row in detail_rows],
        )
    else:
        _write_sheet(detail_sheet, "Detail CF", ["Info"], [["Tidak ada detail perhitungan."]])

    rule_rows = build_rule_rows(results)
    rules_sheet = workbook.create_sheet()
    if rule_rows:
        _write_sheet(
            rules_sheet,
            "Rule FC",
            list(rule_rows[0].keys()),
            [list(row.values()) for row in rule_rows],
        )
    else:
        _write_sheet(rules_sheet, "Rule FC", ["Info"], [["Tidak ada rule yang terpenuhi."]])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
