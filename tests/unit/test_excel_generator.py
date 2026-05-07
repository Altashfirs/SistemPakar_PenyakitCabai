from io import BytesIO

from openpyxl import load_workbook

from engine.cf_engine import forward_chaining
from engine.excel_generator import generate_excel_report


def test_generate_excel_report_contains_expected_sheets_and_summary() -> None:
    user_inputs = {"G01": 1.0, "G02": 0.8, "G03": 0.9, "G25": 0.7}
    results = forward_chaining(user_inputs)

    report_bytes = generate_excel_report(results, user_inputs)
    workbook = load_workbook(BytesIO(report_bytes))

    assert workbook.sheetnames == ["Ringkasan", "Ranking", "Input Gejala", "Detail CF"]

    summary_sheet = workbook["Ringkasan"]
    assert summary_sheet["A2"].value == "Penyakit Utama"
    assert summary_sheet["B2"].value == "Antraknosa (Busuk Buah)"
    assert summary_sheet["A6"].value == "Interpretasi"
    assert summary_sheet["B6"].value == "Sangat Kuat"

    ranking_sheet = workbook["Ranking"]
    assert ranking_sheet["A2"].value == 1
    assert ranking_sheet["B2"].value == "P1"

    inputs_sheet = workbook["Input Gejala"]
    assert inputs_sheet.max_row == 5
    assert inputs_sheet["A2"].value == "G01"

    detail_sheet = workbook["Detail CF"]
    assert detail_sheet["A1"].value == "Kode Penyakit"
    assert detail_sheet.max_row > 1


def test_generate_excel_report_handles_no_selected_symptom_rows() -> None:
    results = forward_chaining({})

    report_bytes = generate_excel_report(results, {})
    workbook = load_workbook(BytesIO(report_bytes))
    inputs_sheet = workbook["Input Gejala"]

    assert inputs_sheet["A2"].value == "-"
    assert inputs_sheet["B2"].value == "Tidak ada gejala dipilih"
    assert inputs_sheet["C2"].value == 0
